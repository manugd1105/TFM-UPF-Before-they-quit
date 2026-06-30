"""
generate_keywords.py
--------------------
Extrae palabras clave de pros y cons del gold.xlsx y las añade
como segunda hoja "keywords" para el word cloud de Tableau.

Uso:
    python generate_keywords.py
"""

import re
from pathlib import Path
import pandas as pd
from collections import Counter
from nltk.corpus import stopwords
from openpyxl import load_workbook

GOLD_PATH = Path(__file__).resolve().parent / "gold.xlsx"
TOP_N     = 100   # top palabras por fuente y categoría de satisfacción

STOP = set(stopwords.words("english")) | {
    # genéricas
    "work", "company", "nvidia", "people", "also", "get", "one",
    "really", "lot", "many", "would", "could", "make", "much",
    "well", "good", "great", "bad", "nothing", "everything",
    "time", "place", "things", "thing", "us", "still", "even",
    "though", "way", "sometimes", "can", "like", "feel", "working",
    "employees", "employee", "team", "teams", "know", "never",
    "always", "ever", "job", "role", "best", "better", "hard",
    # conectores / adjetivos vacíos
    "think", "little", "times", "need", "often", "less", "high",
    "level", "change", "lots", "everyone", "open", "world",
    "also", "building", "companies", "come", "going", "every",
    "almost", "around", "within", "across", "among", "quite",
    "take", "give", "given", "look", "looks", "seem", "seems",
    "maybe", "might", "something", "anything", "nothing",
    # palabras meta o demasiado vagas
    "cons", "pros", "review", "overall", "said", "saying",
    "cutting", "edge",   # solos no aportan; "cutting-edge" sería otro tema
    "free",              # aparece como "free food" pero el contexto se pierde solo
    "long",              # demasiado genérico fuera de "long hours"
}

def tokenize(text):
    if not isinstance(text, str) or not text.strip():
        return []
    text = text.lower()
    text = re.sub(r"[^a-z\s]", " ", text)
    tokens = text.split()
    return [t for t in tokens if t not in STOP and len(t) > 3]

def get_keywords(df, col, label, top_n=TOP_N):
    rows = []
    for satisfaction in ["Alto", "Bajo", "All"]:
        if satisfaction == "All":
            subset = df
        else:
            subset = df[df["satisfaction"] == satisfaction]

        tokens = []
        for text in subset[col].dropna():
            tokens.extend(tokenize(text))

        counter = Counter(tokens)
        for word, freq in counter.most_common(top_n):
            rows.append({
                "word":         word,
                "frequency":    freq,
                "source":       label,        # "pros" o "cons"
                "satisfaction": satisfaction, # "Alto", "Bajo" o "All"
            })
    return rows

# ── Cargar gold.xlsx ──────────────────────────────────────────────────────────
print("Cargando gold.xlsx...")
df = pd.read_excel(GOLD_PATH, sheet_name="gold")
print(f"  {len(df)} filas cargadas")

# ── Extraer keywords ──────────────────────────────────────────────────────────
print("Extrayendo keywords...")
rows = []
rows += get_keywords(df, "pros", "pros")
rows += get_keywords(df, "cons", "cons")

keywords_df = pd.DataFrame(rows)
print(f"  {len(keywords_df)} filas de keywords generadas")
print("\nTop 10 palabras en cons (All):")
sample = keywords_df[(keywords_df["source"]=="cons") & (keywords_df["satisfaction"]=="All")]
print(sample.nlargest(10, "frequency")[["word","frequency"]].to_string(index=False))

# ── Escribir como segunda hoja en gold.xlsx ───────────────────────────────────
print("\nEscribiendo hoja 'keywords' en gold.xlsx...")
wb = load_workbook(GOLD_PATH)

# Eliminar hoja si ya existe
if "keywords" in wb.sheetnames:
    del wb["keywords"]

# Crear hoja nueva
ws = wb.create_sheet("keywords")

# Cabecera
ws.append(["word", "frequency", "source", "satisfaction"])

# Datos
for _, row in keywords_df.iterrows():
    ws.append([row["word"], int(row["frequency"]), row["source"], row["satisfaction"]])

wb.save(GOLD_PATH)
print(f"Hecho. gold.xlsx ahora tiene hojas: {wb.sheetnames}")
