"""
FastAPI — Employee Satisfaction Predictor (NVIDIA Demo)
Pipeline: Tally webhook → n8n → POST /predict → predicción Alto/Bajo
"""

import os
# Debe estar ANTES de cualquier import de numpy/sklearn/xgboost
# Evita SIGSEGV por dos runtimes OpenMP en macOS (XGBoost + sklearn)
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["OMP_NUM_THREADS"] = "1"
import pickle
import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

import nltk
from nltk import word_tokenize, pos_tag
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from nrclex import NRCLex
from empath import Empath
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# Recursos NLTK ya presentes en ~/nltk_data — no se descargan (SSL bloqueado en Mac)

# ── Rutas ────────────────────────────────────────────────────────────────────
# BASE_DIR apunta a la raíz del repo (dos niveles arriba de fastapi/)
BASE_DIR     = Path(__file__).resolve().parent.parent.parent
EMBEDDINGS   = BASE_DIR / "01_outputs_p1/02_Embeddings"
MODEL_PATH   = BASE_DIR / "Dashboard/Modelo FG-1/dict_20260614_144022.pkl"

# ── Constantes NVIDIA (hardcodeadas) ────────────────────────────────────────
NVIDIA_CONSTANTS = {
    # COMPANY — NVIDIA
    "Market Capitalization":   1.069485e+12,
    "NumberofEmployees":       26196,
    "Market Cap per Employee": 4.082628e+07,
    "Founded":                 1993,
    # TARGET_ENC — Electronic Technology sector (valores del train set)
    "Sector_targenc_median":   4.0,
    "Sector_targenc_mean":     3.98,
}

# SECTOR OHE — Electronic Technology = 1, resto = 0
ALL_SECTORS = [
    "Commercial Services","Communications","Consumer Durables","Consumer Non-Durables",
    "Consumer Services","Distribution Services","Electronic Technology","Energy Minerals",
    "Finance","Health Services","Health Technology","Industrial Services",
    "Non-Energy Minerals","Process Industries","Producer Manufacturing","Retail Trade",
    "Technology Services","Transportation","Utilities"
]

# Empath categories usadas en el modelo
EMPATH_CATS = [
    "help","office","money","domestic_work","sleep","occupation","family",
    "vacation","health","swearing_terms","leisure","suffering","wealthy",
    "exercise","home","rage","fun","negative_emotion","payment","achievement"
]

# ODI vocabulary reconstruido (keyword matching de las 9 dimensiones clínicas)
ODI_VOCAB = {
    "stress","stressed","anxiety","anxious","worry","worried","fear","dread",
    "depressed","depression","sad","sadness","unhappy","hopeless","despair",
    "burnout","burn out","exhausted","exhaustion","fatigue","tired","drained",
    "overwhelmed","sleep","insomnia","rest","restless","appetite","eating",
    "worthless","worthlessness","useless","failure","inadequate","inferior",
    "anhedonia","joy","pleasure","enjoy","enjoyment","motivation","interest",
    "cognitive","focus","concentrate","concentration","memory","distracted",
    "psychomotor","slow","sluggish","agitated","restless","lethargic",
    "suicidal","suicide","self-harm","harm","death","die","hopeless",
    "mental","mental health","wellbeing","well-being","morale","toxic",
    "pressure","frustration","frustrated","irritable","irritability",
}

# JDI dimensiones (orden = fila en xlsx)
JDI_DIMS = ["Work itself", "Pay", "Promotion", "Supervision", "Coworkers"]
# ODI dimensiones (orden = fila en xlsx)
ODI_DIMS = [
    "Depressed mood","Sleep alterations","Fatigue","Worthlessness","Anhedonia",
    "Appetite alterations","Cognitive impairment","Psycomotor alterations","Suicidal ideation"
]

# Fecha t0 de referencia del train set (última fecha observada)
T0 = datetime.date(2023, 12, 31)
NVIDIA_FOUNDED = datetime.date(1993, 1, 5)

# ── Startup: carga de modelos y embeddings ───────────────────────────────────
app = FastAPI(title="NVIDIA Satisfaction Predictor", version="1.0")

print("🔄 Cargando modelo XGBoost...")
with open(MODEL_PATH, "rb") as f:
    _pkl = pickle.load(f)
MODEL = _pkl["Random search object"].best_estimator_

print("🔄 Cargando feature order desde el modelo...")
FEATURE_ORDER = list(MODEL.feature_names_in_)

print("🔄 Cargando embeddings JDI/ODI...")
JDI_POS = pd.read_excel(EMBEDDINGS / "dimensions_JDI_pos_raw.xlsx", header=0).select_dtypes(include=[np.number]).values
JDI_NEG = pd.read_excel(EMBEDDINGS / "dimensions_JDI_neg_raw.xlsx", header=0).select_dtypes(include=[np.number]).values
ODI_POS = pd.read_excel(EMBEDDINGS / "dimensions_ODI_pos_raw.xlsx", header=0).select_dtypes(include=[np.number]).values
ODI_NEG = pd.read_excel(EMBEDDINGS / "dimensions_ODI_neg_raw.xlsx", header=0).select_dtypes(include=[np.number]).values

print("🔄 Cargando sentence-transformer (all-roberta-large-v1)...")
ST_MODEL = SentenceTransformer("all-roberta-large-v1")

print("🔄 Inicializando VADER y Empath...")
SIA     = SentimentIntensityAnalyzer()
EMPATH  = Empath()

print("✅ Todo cargado. FastAPI listo.")


# ── Input schema ─────────────────────────────────────────────────────────────
class ReviewInput(BaseModel):
    summary:    str
    pros:       str
    cons:       str
    department: str = "Engineering & Tech"
    seniority:  str = "Senior"
    gender:     str = "Prefer not to say"
    wl_balance:            float = 3.0
    culture_values:        float = 3.0
    diversity_inclusion:   float = 3.0
    career_opportunities:  float = 3.0
    compensation_benefits: float = 3.0
    senior_management:     float = 3.0
    submission_date: str = ""   # ISO format YYYY-MM-DD, opcional


# ── Helpers de feature engineering ──────────────────────────────────────────

def pos_features(text: str, prefix: str) -> dict:
    """Frecuencias relativas de POS tags."""
    tokens = word_tokenize(text)
    total  = len(tokens) if tokens else 1
    tags   = pos_tag(tokens)
    freq   = {}
    for _, tag in tags:
        freq[tag] = freq.get(tag, 0) + 1
    return {f"{prefix}_ld_{tag}": count / total for tag, count in freq.items()}


def emolex_features(text: str, prefix: str) -> dict:
    """EmoLex affect frequencies. anticipation siempre 0 (bug original)."""
    lex   = NRCLex(text)
    freqs = lex.affect_frequencies
    cats  = ["fear","anger","anticip","trust","surprise","positive","negative",
             "sadness","disgust","joy"]
    result = {}
    for cat in cats:
        key = f"{prefix}_emolex_{cat}"
        if cat == "anticip":
            result[key] = 0.0          # replica bug versión original
        else:
            # NRCLex usa nombres distintos internamente
            nrc_key = {
                "fear":"fear","anger":"anger","trust":"trust",
                "surprise":"surprise","positive":"positive","negative":"negative",
                "sadness":"sadness","disgust":"disgust","joy":"joy"
            }.get(cat, cat)
            result[key] = freqs.get(nrc_key, 0.0)
    return result


def empath_features(text: str, prefix: str) -> dict:
    """Empath normalized scores."""
    result = EMPATH.analyze(text, categories=EMPATH_CATS, normalize=True) or {}
    return {f"{prefix}_empath_{cat}": result.get(cat, 0.0) for cat in EMPATH_CATS}


def vader_feature(text: str, prefix: str) -> dict:
    """VADER compound score."""
    return {f"{prefix}_nltk_sia": SIA.polarity_scores(text)["compound"]}


def odi_vocab_feature(text: str, prefix: str) -> dict:
    """Proporción de palabras del texto que están en el ODI vocabulary."""
    words = text.lower().split()
    total = len(words) if words else 1
    count = sum(1 for w in words if w in ODI_VOCAB)
    return {f"ODI_related_vocab_{prefix.split('_')[0]}": count / total}


def length_features(text: str, prefix: str) -> dict:
    return {
        f"{prefix}_charlen": len(text),
        f"{prefix}_wordlen": len(text.split()),
    }


def embedding_features(text: str, prefix: str, dim_matrix_pos, dim_matrix_neg,
                       dim_names: list, sim_prefix: str) -> dict:
    """Cosine similarity entre embedding del texto y cada dimensión JDI/ODI."""
    emb = ST_MODEL.encode(text)  # (1024,)
    emb = emb.reshape(1, -1)

    # pros → pos file, cons → neg file
    if "pros" in prefix:
        matrix = dim_matrix_pos
    else:
        matrix = dim_matrix_neg

    result = {}
    for i, dim_name in enumerate(dim_names):
        dim_emb = matrix[i].reshape(1, -1)
        sim     = cosine_similarity(emb, dim_emb)[0][0]
        result[f"{prefix}_sim_{sim_prefix}_{dim_name}"] = float(sim)
    return result


def get_stock_features(review_date: datetime.date) -> dict:
    """Stock % change para NVDA y SP500 en el mes y año de la review."""
    try:
        # Mes anterior al de la review
        start_month = (review_date.replace(day=1) - datetime.timedelta(days=32)).replace(day=1)
        end_month   = review_date

        nvda = yf.download("NVDA", start=start_month, end=end_month,
                           progress=False, auto_adjust=True)
        sp   = yf.download("SPY", start=start_month, end=end_month,
                           progress=False, auto_adjust=True)

        def pct_change(df):
            if df.empty or len(df) < 2:
                return 0.0
            return float((df["Close"].iloc[-1] - df["Close"].iloc[0]) / df["Close"].iloc[0])

        # Año completo
        start_year = review_date.replace(month=1, day=1)
        nvda_y = yf.download("NVDA", start=start_year, end=end_month,
                              progress=False, auto_adjust=True)
        sp_y   = yf.download("SPY", start=start_year, end=end_month,
                              progress=False, auto_adjust=True)

        return {
            "stock_percchange_month": pct_change(nvda),
            "stock_percchange_year":  pct_change(nvda_y),
            "sp500_percchange_month": pct_change(sp),
            "sp500_percchange_year":  pct_change(sp_y),
        }
    except Exception:
        # Si falla yfinance, valores neutros
        return {
            "stock_percchange_month": 0.0,
            "stock_percchange_year":  0.0,
            "sp500_percchange_month": 0.0,
            "sp500_percchange_year":  0.0,
        }


def temporal_features(review_date: datetime.date) -> dict:
    """Features temporales calculadas respecto a T0."""
    days_review_to_t0     = (T0 - review_date).days
    days_foundation_to_t0 = (T0 - NVIDIA_FOUNDED).days
    days_foundation_to_review = (review_date - NVIDIA_FOUNDED).days
    return {
        "month":                          review_date.month,
        "year":                           review_date.year,
        "days from review to t0":         days_review_to_t0,
        "days from foundation to t0":     days_foundation_to_t0,
        "days from foundation to review": days_foundation_to_review,
    }


def sector_ohe_features() -> dict:
    """Sector OHE — Electronic Technology = 1, resto = 0."""
    result = {}
    for sector in ALL_SECTORS:
        key = f"Sector_{sector}_ohe"
        result[key] = 1 if sector == "Electronic Technology" else 0
    return result


# ── Endpoint principal ───────────────────────────────────────────────────────

@app.post("/predict")
def predict(review: ReviewInput):
    # Fecha de la review
    if review.submission_date:
        try:
            review_date = datetime.date.fromisoformat(review.submission_date)
        except ValueError:
            review_date = datetime.date.today()
    else:
        review_date = datetime.date.today()

    # ── Feature engineering ──────────────────────────────────────────────────
    features = {}

    # 1. Lengths
    features.update(length_features(review.summary, "summary"))
    features.update(length_features(review.pros,    "pros"))
    features.update(length_features(review.cons,    "cons"))

    # 2. POS tags
    features.update(pos_features(review.summary, "summary"))
    features.update(pos_features(review.pros,    "pros"))
    features.update(pos_features(review.cons,    "cons"))

    # 3. EmoLex
    features.update(emolex_features(review.pros, "pros"))
    features.update(emolex_features(review.cons, "cons"))

    # 4. VADER
    features.update(vader_feature(review.pros, "pros"))
    features.update(vader_feature(review.cons, "cons"))

    # 5. ODI vocab
    features.update(odi_vocab_feature(review.pros, "pros"))
    features.update(odi_vocab_feature(review.cons, "cons"))

    # 6. Empath
    features.update(empath_features(review.pros, "pros"))
    features.update(empath_features(review.cons, "cons"))

    # 7. JDI similarities
    features.update(embedding_features(
        review.pros, "pros", JDI_POS, JDI_NEG, JDI_DIMS, "JDI"))
    features.update(embedding_features(
        review.cons, "cons", JDI_POS, JDI_NEG, JDI_DIMS, "JDI"))

    # 8. ODI similarities
    features.update(embedding_features(
        review.pros, "pros", ODI_POS, ODI_NEG, ODI_DIMS, "ODI"))
    features.update(embedding_features(
        review.cons, "cons", ODI_POS, ODI_NEG, ODI_DIMS, "ODI"))

    # 9. NVIDIA constants (company + sector target encoding)
    features.update(NVIDIA_CONSTANTS)

    # 10. Sector OHE
    features.update(sector_ohe_features())

    # 12. Stock features
    features.update(get_stock_features(review_date))

    # 13. Temporal
    features.update(temporal_features(review_date))

    # ── Construir vector en el orden exacto del modelo ───────────────────────
    vector = []
    missing = []
    for feat in FEATURE_ORDER:
        val = features.get(feat, None)
        if val is None:
            missing.append(feat)
            vector.append(0.0)   # default 0 para features POS que no aparecen en este texto
        else:
            vector.append(float(val))

    X = np.array(vector).reshape(1, -1)

    # ── Predicción ───────────────────────────────────────────────────────────
    pred_raw = float(MODEL.predict(X)[0])   # XGBRegressor → float 1-5

    # Binarización: >= 4 → Alto, < 4 → Bajo
    satisfaction = "Alto" if pred_raw >= 4.0 else "Bajo"

    # Proxy de confianza: distancia del umbral (4.0) normalizada a [0, 1]
    confidence = round(min(abs(pred_raw - 4.0) / 2.0, 1.0), 3)

    return {
        "satisfaction":    satisfaction,
        "predicted_rating": float(pred_raw),
        "confidence":      round(confidence, 3),
        "department":      review.department,
        "seniority":       review.seniority,
        "gender":          review.gender,
        "review_date":     review_date.isoformat(),
        "missing_features": missing[:10] if missing else [],  # primeros 10 para debug
    }


@app.get("/health")
def health():
    return {"status": "ok", "model": str(MODEL_PATH.name)}


# ── Append gold endpoint ──────────────────────────────────────────────────────
GOLD_XLSX = BASE_DIR / "Dashboard/gold.xlsx"
GOLD_COLS = [
    "timestamp", "submission_date", "summary", "pros", "cons",
    "department", "seniority", "gender",
    "wl_balance", "culture_values", "diversity_inclusion",
    "career_opportunities", "compensation_benefits", "senior_management",
    "predicted_rating", "satisfaction", "confidence",
]

class GoldRow(BaseModel):
    timestamp:              str   = ""
    submission_date:        str   = ""
    summary:                str   = ""
    pros:                   str   = ""
    cons:                   str   = ""
    department:             str   = ""
    seniority:              str   = ""
    gender:                 str   = ""
    wl_balance:             float = None
    culture_values:         float = None
    diversity_inclusion:    float = None
    career_opportunities:   float = None
    compensation_benefits:  float = None
    senior_management:      float = None
    predicted_rating:       float = None
    satisfaction:           str   = ""
    confidence:             float = None

@app.post("/append-gold")
def append_gold(row: GoldRow):
    from openpyxl import load_workbook
    wb = load_workbook(GOLD_XLSX)
    ws = wb.active
    values = [getattr(row, col) for col in GOLD_COLS]
    ws.append(values)
    wb.save(GOLD_XLSX)
    return {"status": "ok", "rows": ws.max_row - 1}
